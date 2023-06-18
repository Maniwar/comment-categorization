import streamlit as st
import pandas as pd
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.sentiment import SentimentIntensityAnalyzer
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import base64
from io import BytesIO
import matplotlib.pyplot as plt
import seaborn as sns
import datetime
import numpy as np
import openpyxl

# Specify download directory for NLTK data
nltk.download('stopwords', download_dir='/home/appuser/nltk_data')
nltk.download('vader_lexicon', download_dir='/home/appuser/nltk_data')
nltk.download('punkt', download_dir='/home/appuser/nltk_data', quiet=True)  # Add 'quiet=True' to suppress NLTK download messages
nltk.data.path.append('/home/appuser/nltk_data')

# Initialize BERT model
model = SentenceTransformer('all-MiniLM-L6-v2')

# Create a dictionary to store precomputed embeddings
keyword_embeddings = {}

# Preprocessing function
def preprocess_text(text):
    # Remove unnecessary characters
    text = str(text).strip().lower()

    # Remove stopwords
    stop_words = set(stopwords.words('english'))
    word_tokens = word_tokenize(text)
    filtered_text = [word for word in word_tokens if word.lower() not in stop_words]
    preprocessed_text = ' '.join(filtered_text)

    return preprocessed_text

# Perform sentiment analysis
def perform_sentiment_analysis(text):
    sia = SentimentIntensityAnalyzer()
    sentiment_scores = sia.polarity_scores(text)
    compound_score = sentiment_scores['compound']
    return compound_score

# Compute semantic similarity
def compute_semantic_similarity(keyword_embedding, comment_embedding):
    similarity = cosine_similarity([keyword_embedding], [comment_embedding])
    return similarity[0][0]

# Streamlit interface
st.title("Feedback Categorization")

# File upload
uploaded_file = st.file_uploader("Upload CSV file", type="csv")

# Select the column containing the comments
comment_column = None
date_column = None
if uploaded_file is not None:
    # Read customer feedback from uploaded file
    csv_data = uploaded_file.read()
    feedback_data = pd.read_csv(BytesIO(csv_data))
    comment_column = st.selectbox("Select the column containing the comments", feedback_data.columns.tolist())
    date_column = st.selectbox("Select the column containing the dates", feedback_data.columns.tolist())
    grouping_option = st.radio("Select how to group the dates", ["Date", "Week", "Month", "Quarter"])
    process_button = st.button("Process Feedback")

    if comment_column is not None and date_column is not None and grouping_option is not None and process_button:
        # Define keyword-based categories
        default_categories = {
            'Website Usability': ['user experience', 'navigation', 'interface', 'design', 'loading speed', 'search functionality', 'mobile responsiveness', 'compatibility', 'site organization', 'site performance', 'page loading time', 'difficult to navigate', 'poorly organized', 'difficult to find information', 'not user-friendly', 'hard to use'],
            'Product Quality': ['quality', 'defects', 'durability', 'reliability', 'performance', 'features', 'accuracy', 'packaging', 'product condition'],
            'Shipping and Delivery': ['delivery', 'shipping', 'shipping time', 'tracking', 'package condition', 'courier', 'fulfillment', 'damaged during shipping', 'order tracking'],
            'Customer Support': ['support service', 'support response time', 'communication', 'issue resolution', 'knowledgeability', 'helpfulness', 'replacement', 'refund', 'customer service experience'],
            'Order Processing': ['payment', 'stock availability', 'order confirmation', 'cancellation', 'order customization', 'order accuracy'],
            'Product Selection': ['product customization', 'customization difficulties', 'product variety', 'product options', 'finding the right product', 'product suitability'],
            'Price Change in Cart': ['promotions', 'price discrepancy', 'discounts missing', 'rebate missing', 'price change in cart', 'price change at checkout', 'trade-in discount', 'free memory upgrade', 'discount program', 'EPP', 'employee discount'],
            'Product Information': ['specifications', 'descriptions', 'images', 'product details', 'accurate information', 'misleading information', 'product data'],
            'Returns and Refunds': ['returns', 'refunds', 'return policy', 'refund process', 'return condition', 'return shipping', 'return authorization'],
            'Warranty and Support': ['warranty', 'technical support', 'repair', 'technical assistance', 'warranty claim', 'product support'],
            'Website Security': ['security', 'privacy', 'data protection', 'secure checkout', 'account security', 'payment security'],
            'Website Performance': ['uptime', 'site speed', 'loading time', 'server stability', 'site crashes', 'website availability'],
            'Accessibility': ['accessibility', 'inclusive design', 'special needs', 'assistive technologies', 'screen reader compatibility', 'website usability for disabled users'],
            'Unwanted Emails': ['spam emails', 'email subscriptions', 'unsubscribe', 'email preferences', 'inbox management', 'email marketing'],
        }

        # Edit categories and keywords
        st.sidebar.header("Edit Categories")
        categories = {}
        for category, keywords in default_categories.items():
            category_name = st.sidebar.text_input(f"{category} Category", value=category)
            category_keywords = st.sidebar.text_area(f"Keywords for {category}", value="\n".join(keywords))
            categories[category_name] = category_keywords.split("\n")

        st.sidebar.subheader("Add or Modify Categories")
        new_category_name = st.sidebar.text_input("New Category Name")
        new_category_keywords = st.sidebar.text_area(f"Keywords for {new_category_name}")
        if new_category_name and new_category_keywords:
            categories[new_category_name] = new_category_keywords.split("\n")

        # Initialize lists to store categorized comments and sentiments
        categorized_comments = []
        sentiments = []

        # Pre-compute embeddings for all keywords
        for category_name, keywords in categories.items():
            for keyword in keywords:
                keyword_embeddings[keyword] = model.encode([keyword])[0]

        # Process each comment
        with st.spinner('Processing feedback...'):
            for index, row in feedback_data.iterrows():
                preprocessed_comment = preprocess_text(row[comment_column])
                comment_embedding = model.encode([preprocessed_comment])[0]  # Compute the comment embedding once
                sentiment_score = perform_sentiment_analysis(preprocessed_comment)

                category = 'Other'
                sub_category = 'Other'
                best_match_score = float('-inf')  # Initialized to negative infinity

                # Tokenize the preprocessed_comment
                tokens = word_tokenize(preprocessed_comment)

                for main_category, keywords in categories.items():
                    for keyword in keywords:
                        keyword_embedding = keyword_embeddings[keyword]  # Use the precomputed keyword embedding
                        similarity_score = compute_semantic_similarity(keyword_embedding, comment_embedding)
                        # If similarity_score equals best_match_score, we pick the first match.
                        # If similarity_score > best_match_score, we update best_match.
                        if similarity_score >= best_match_score:
                            category = main_category
                            sub_category = keyword
                            best_match_score = similarity_score

                parsed_date = row[date_column].split(' ')[0] if isinstance(row[date_column], str) else None
                row_extended = row.tolist() + [preprocessed_comment, category, sub_category, sentiment_score, parsed_date]
                categorized_comments.append(row_extended)
                sentiments.append(sentiment_score)

        # Generate trends and insights
        st.success('Done!')
        existing_columns = feedback_data.columns.tolist()
        additional_columns = [comment_column, 'Category', 'Sub-Category', 'Sentiment', 'Parsed Date']
        headers = existing_columns + additional_columns

        # Create a new DataFrame with extended columns
        trends_data = pd.DataFrame(categorized_comments, columns=headers)
        trends_data['Parsed Date'] = pd.to_datetime(trends_data['Parsed Date'], errors='coerce').dt.date


        # Rename duplicate column names
        trends_data = trends_data.loc[:, ~trends_data.columns.duplicated()]
        duplicate_columns = set([col for col in trends_data.columns if trends_data.columns.tolist().count(col) > 1])
        for column in duplicate_columns:
            column_indices = [i for i, col in enumerate(trends_data.columns) if col == column]
            for i, idx in enumerate(column_indices[1:], start=1):
                trends_data.columns.values[idx] = f"{column}_{i}"

        # Display trends and insights
        st.title("Feedback Trends and Insights")
        st.dataframe(trends_data)

        # Display pivot table with counts for Category, Sub-Category, and Parsed Date
        st.subheader("All Categories Trends")

        # Convert 'Parsed Date' into datetime format if it's not
        trends_data['Parsed Date'] = pd.to_datetime(trends_data['Parsed Date'], errors='coerce')

        # Create pivot table with counts for Category, Sub-Category, and Parsed Date
        if grouping_option == 'Date':
            pivot = trends_data.pivot_table(
                index=['Category', 'Sub-Category'],
                columns=pd.Grouper(key='Parsed Date', freq='D'),
                values='Sentiment',
                aggfunc='count',
                fill_value=0
            )

        elif grouping_option == 'Week':
            pivot = trends_data.pivot_table(
                index=['Category', 'Sub-Category'],
                columns=pd.Grouper(key='Parsed Date', freq='W-SUN'),  # Adjusted to start week on Sunday
                values='Sentiment',
                aggfunc='count',
                fill_value=0
            )
            # Shift the pivot table columns by -1 week
            pivot.columns = pivot.columns - pd.DateOffset(weeks=1)

        elif grouping_option == 'Month':
            pivot = trends_data.pivot_table(
                index=['Category', 'Sub-Category'],
                columns=pd.Grouper(key='Parsed Date', freq='M'),
                values='Sentiment',
                aggfunc='count',
                fill_value=0
            )
        elif grouping_option == 'Quarter':
            pivot = trends_data.pivot_table(
                index=['Category', 'Sub-Category'],
                columns=pd.Grouper(key='Parsed Date', freq='Q'),
                values='Sentiment',
                aggfunc='count',
                fill_value=0
            )

        pivot.columns = pivot.columns.astype(str)  # Convert column labels to strings

        # Sort the pivot table rows based on the highest count
        pivot = pivot.loc[pivot.sum(axis=1).sort_values(ascending=False).index]

        # Sort the pivot table columns in descending order based on the most recent date
        pivot = pivot[sorted(pivot.columns, reverse=True)]

        # Create a line chart for the top 5 trends over time with the selected grouping option
        plt.figure(figsize=(10, 6))

        for sub_category in pivot.head(5).index:
            # Get the dates with non-zero counts for the sub-category and the selected grouping option
            non_zero_dates = pivot.columns[(pivot.loc[sub_category] > 0) & (pivot.columns.isin(pivot.columns))]
            # Get the corresponding data points for the non-zero dates
            non_zero_counts = pivot.loc[sub_category, non_zero_dates]
            # Plot the line for the sub-category
            plt.plot(non_zero_dates, non_zero_counts, marker='o', label=sub_category)

        plt.xlabel('Date')
        plt.ylabel('Count')
        plt.title('Top 5 Trends Over Time')

        # Set the x-axis ticks and labels based on the selected grouping option
        if grouping_option == 'Date':
            plt.xticks(rotation=45)

        elif grouping_option == 'Week':
            week_dates = pd.date_range(pivot.columns.min(), pivot.columns.max(), freq='W-SUN')
            week_dates = week_dates[week_dates.isin(pivot.columns)]
            week_numbers = [date.strftime('%Y-%m-%d') for date in week_dates]
            week_dates_str = [date.strftime('%Y-%m-%d') for date in week_dates]
            plt.xticks(week_dates_str, week_numbers, rotation=45)

        elif grouping_option == 'Month':
            month_dates = pd.date_range(pivot.columns.min(), pivot.columns.max(), freq='M')
            month_dates = month_dates[month_dates.isin(pivot.columns)]
            month_numbers = [date.strftime('%Y-%m') for date in month_dates]
            month_dates_str = [date.strftime('%Y-%m-%d') for date in month_dates]
            plt.xticks(month_dates_str, month_numbers, rotation=45)

        elif grouping_option == 'Quarter':
            quarter_dates = pd.date_range(pivot.columns.min(), pivot.columns.max(), freq='Q')
            quarter_dates = quarter_dates[quarter_dates.isin(pivot.columns)]
            quarter_labels = [f'Q{date.quarter}\n{date.year}' for date in quarter_dates]
            quarter_dates_str = [date.strftime('%Y-%m-%d') for date in quarter_dates]
            plt.xticks(quarter_dates_str, quarter_labels, rotation=45)

        plt.legend()
        plt.tight_layout()

        # Display the line chart
        st.pyplot(plt.gcf())

        # Display pivot table with counts for Category, Sub-Category, and Parsed Date
        st.dataframe(pivot)

        # Create pivot tables with counts
        pivot1 = trends_data.pivot_table(index='Category', values='Sentiment', aggfunc=['mean', 'count'])
        pivot1.columns = ['Average Sentiment', 'Survey Count']
        pivot1 = pivot1.sort_values('Survey Count', ascending=False)

        pivot2 = trends_data.pivot_table(index=['Category', 'Sub-Category'], values='Sentiment', aggfunc=['mean', 'count'])
        pivot2.columns = ['Average Sentiment', 'Survey Count']
        pivot2 = pivot2.sort_values('Survey Count', ascending=False)

        # Create and display a bar chart for pivot1 with counts
        plt.figure(figsize=(10, 6))
        sns.barplot(x=pivot1.index, y=pivot1['Survey Count'])
        plt.title("Survey Count by Category")
        plt.xticks(rotation=90)
        plt.tight_layout()
        st.pyplot(plt.gcf())

        # Display pivot table with counts for Category
        st.subheader("Category vs Sentiment and Survey Count")
        st.dataframe(pivot1)

        # Create and display a bar chart for pivot2 with counts
        plt.figure(figsize=(10, 6))
        sns.barplot(x=pivot2.index.get_level_values(1), y=pivot2['Survey Count'])
        plt.title("Survey Count by Sub-Category")
        plt.xticks(rotation=90)
        plt.tight_layout()
        st.pyplot(plt.gcf())

        # Display pivot table with counts for Sub-Category
        st.subheader("Sub-Category vs Sentiment and Survey Count")
        st.dataframe(pivot2)


        # Format 'Parsed Date' as string with 'YYYY-MM-DD' format
        trends_data['Parsed Date'] = trends_data['Parsed Date'].dt.strftime('%Y-%m-%d').fillna('')

        # Create pivot table with counts for Category, Sub-Category, and Parsed Date
        pivot = trends_data.pivot_table(
            index=['Category', 'Sub-Category'],
            columns=pd.to_datetime(trends_data['Parsed Date']).dt.strftime('%Y-%m-%d'),  # Format column headers as 'YYYY-MM-DD'
            values='Sentiment',
            aggfunc='count',
            fill_value=0
        )

        # Sort the pivot table rows based on the highest count
        pivot = pivot.loc[pivot.sum(axis=1).sort_values(ascending=False).index]

        # Sort the pivot table columns in descending order based on the most recent date
        pivot = pivot[sorted(pivot.columns, reverse=True)]

        # Save DataFrame and pivot tables to Excel
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            trends_data.to_excel(writer, sheet_name='Feedback Trends and Insights', index=False)

            # Convert 'Parsed Date' column to datetime type
            trends_data['Parsed Date'] = pd.to_datetime(trends_data['Parsed Date'], errors='coerce')

            # Create a separate column for formatted date strings
            trends_data['Formatted Date'] = trends_data['Parsed Date'].dt.strftime('%Y-%m-%d')

            # Reset the index
            trends_data.reset_index(inplace=True)

            # Set 'Formatted Date' column as the index
            trends_data.set_index('Formatted Date', inplace=True)

            # Create pivot table with counts for Category, Sub-Category, and Parsed Date
            if grouping_option == 'Date':
                pivot = trends_data.pivot_table(
                    index=['Category', 'Sub-Category'],
                    columns='Parsed Date',  # Use the column name directly
                    values='Sentiment',
                    aggfunc='count',
                    fill_value=0
                )
            elif grouping_option == 'Week':
                pivot = trends_data.pivot_table(
                    index=['Category', 'Sub-Category'],
                    columns=pd.Grouper(key='Parsed Date', freq='W-SUN'),
                    values='Sentiment',
                    aggfunc='count',
                    fill_value=0
                )
                # Shift the pivot table columns by -1 week
                pivot.columns = pivot.columns - pd.DateOffset(weeks=1)
            elif grouping_option == 'Month':
                pivot = trends_data.pivot_table(
                    index=['Category', 'Sub-Category'],
                    columns=pd.Grouper(key='Parsed Date', freq='M'),
                    values='Sentiment',
                    aggfunc='count',
                    fill_value=0
                )
            elif grouping_option == 'Quarter':
                pivot = trends_data.pivot_table(
                    index=['Category', 'Sub-Category'],
                    columns=pd.Grouper(key='Parsed Date', freq='Q'),
                    values='Sentiment',
                    aggfunc='count',
                    fill_value=0
                )

            # Format column headers as date strings in 'YYYY-MM-DD' format
            pivot.columns = pivot.columns.strftime('%Y-%m-%d')


            pivot.to_excel(writer, sheet_name='Trends by ' + grouping_option, merge_cells=False)

            pivot1.to_excel(writer, sheet_name='Categories', merge_cells=False)
            pivot2.to_excel(writer, sheet_name='Subcategories', merge_cells=False)


        excel_file.seek(0)
        b64 = base64.b64encode(excel_file.read()).decode()
        href = f'<a href="data:application/octet-stream;base64,{b64}" download="feedback_trends.xlsx">Download Excel File</a>'
        st.markdown(href, unsafe_allow_html=True)
